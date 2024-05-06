import openpyxl

excel = openpyxl.Workbook()
ws = excel.active

ws.title = "db_alunos"
db_alunos = excel['db_alunos']

# Adiciona aluno na db -> atualmente é manual. Precisa de uma tela para ficar mais elegante.
def add_aluno(nome, matricula):
    db_alunos.append([f'{nome}', f'{matricula}'])

# Pega as informações do aluno por matrícula ou nome
def info_aluno(nome='', matricula=0):
    for i in range (db_alunos.max_row):
        if nome == db_alunos.cell(i+1,1).value or str(matricula) == db_alunos.cell(i+1,2).value:

            nome_aluno = db_alunos.cell(i+1,1).value
            matricula_aluno = db_alunos.cell(i+1,2).value
            linha_infoaluno = i+1
            print(f'Linha {i+1} - Aluno {nome_aluno} - Matrícula {matricula_aluno}')
            return nome_aluno, matricula_aluno, linha_infoaluno


# Testes para verificar se está batendo.
# add_aluno('hugo', '1234567891')
# nome_aluno, matricula_aluno, linha_infoaluno = info_aluno(nome='' ,matricula=1234567891)
# nome_aluno, matricula_aluno, linha_infoaluno = info_aluno(nome='hugo' ,matricula=0)